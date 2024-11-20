# Report_Generator.py
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
from datetime import datetime
from openpyxl.utils import get_column_letter

class ReportGenerator:
    def __init__(self, filename):
        self.filename = filename
        self.passed_tests = []
        self.failed_tests = []

    def add_test_result(self, test_name, status, message):
        if status == "Online":
            self.passed_tests.append((test_name, message))
        else:
            self.failed_tests.append((test_name, message))

    def generate_report(self):
        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        report_name = f"{self.filename}_{now}.xlsx"
        self.save_excel_report(report_name)
        return report_name



    def save_excel_report(self, report_name):
        try:
            # Ensure the directory exists
            output_directory = os.getcwd()
            full_path = os.path.join(output_directory, report_name)

            # Create a new Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "FleetMGT Daily Report"

            # Add a summary section
            total_tests = len(self.passed_tests) + len(self.failed_tests)
            passed_tests = len(self.passed_tests)
            failed_tests = len(self.failed_tests)

            ws.append(["Total Device Count", total_tests])
            ws.append(["Online Devices Count", passed_tests])
            ws.append(["Offline Devices Count", failed_tests])
            ws.append([])  # Blank row for separation

            # Add test details
            ws.append(["Device List", "Status", "Last Updated Date"])
            pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Write the passed test cases
            for test_name, message in self.passed_tests:
                ws.append([test_name, "Online", message])
                for cell in ws[ws.max_row]:
                    cell.fill = pass_fill

            # Write the failed test cases
            for test_name, message in self.failed_tests:
                ws.append([test_name, "Offline", message])
                for cell in ws[ws.max_row]:
                    cell.fill = fail_fill

            # Set column widths dynamically based on the content
            for col in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in ws.iter_rows(min_col=col, max_col=col):
                    for cell in row:
                        try:
                            # Get the length of the string value in the cell
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                adjusted_width = (max_length + 2)  # Add some padding
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            wb.save(full_path)
            print(f"Report successfully saved: {full_path}")
            return full_path

        except Exception as e:
            print(f"Error while saving report: {e}")
            raise

    def send_report_via_email(self, email_list, app_password, sender_email, report_name):
        try:
            # Set up the email server
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(sender_email, app_password)

            for recipient_email in email_list:
                msg = MIMEMultipart()
                msg['From'] = sender_email
                msg['To'] = recipient_email
                msg['Subject'] = f"FleetMGT Daily Report"

                body = "Please find the attached test report."
                msg.attach(MIMEText(body, 'plain'))

                # Attach the Excel report
                with open(report_name, 'rb') as report_file:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(report_file.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', f'attachment; filename={report_name}')
                    msg.attach(part)

                # Send the email
                server.sendmail(sender_email, recipient_email, msg.as_string())
                print(f"Report sent to {recipient_email}")

        except Exception as e:
            print(f"Error sending email: {e}")
        finally:
            server.quit()
